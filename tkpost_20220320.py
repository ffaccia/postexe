
import requests
import sys
import os
import json
from shutil import copyfile as sh_copy
import base64
import babel
import babel.numbers

from datetime import datetime as dt
import time
from dotenv import load_dotenv

import sqlite3
import logging

from tkinter import *
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
from tkcalendar import DateEntry

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter

from PIL import ImageTk, Image

from utils import round_format_size as rs

FORMAT = "[ %(asctime)s, %(levelname)s] %(message)s" 



load_dotenv()
CONFIGURATION_FILE = os.getenv('CONFIGURATION_FILE')

                 
dbFile, conn, cur = (None,)*3
entry_dir, response_dir, save_dir, export_dir = (None,)*4
logs_dir, db_dir, img_dir, data_dir = (None,)*4
CODES_OK = None

global data

im_checked, im_unchecked, trv = (None,)*3
checkall_btn, uncheckall_btn, upload_again_btn, excel_btn = (None,)*4
allowedFileTypes=(('pdf files', '*.pdf'),)

SAVE_FILE_DB = True   
SAVE_FILE_DIR = True   
W_WIDTH = 900
W_HEIGHT = 600

root = Tk()
root.title("PDF 2 WS")
root.geometry("%dx%d" % (W_WIDTH, W_HEIGHT))
   

def load_config(file_config):
    global data
    with open(file_config, "r") as jsonfile:
        data = json.load(jsonfile)
        DATA_DIR = data['DATA_DIR']

        
def setup_profile():    
    global CODES_OK
    global im_checked, im_unchecked
    global response_dir, save_dir, export_dir
    global logs_dir, db_dir, img_dir, data_dir
    
    
    #ff, modified at 20220328
    #load_config("test_config.json")
    load_config(CONFIGURATION_FILE)
    
    
    CODES_OK = [ v for k,v in requests.codes.__dict__.items() if k in data['CODES_OK']]
    #MULTI_FILES = True

    try:
        user_profile_dir = os.getcwd()
        if user_profile_dir and os.path.isdir(user_profile_dir) == True:
            
            try:
                entry_dir = user_profile_dir
                #inner_dir = os.path.join('documents','sviluppo','python','postexe')
                if os.path.isdir(entry_dir) == False: 
                    print("innerdir")
                    print(entry_dir)
                    os.chdir(user_profile_dir)  
                    os.makedirs(entry_dir)
            
            except:    
                msg = "Setup dir is missing: %s! " % (entry_dir)
                tk.messagebox.showerror("Setup Error", msg)
                raise IOError("Errore Creazione Directory", msg)    
                
        else:
                msg = "User Profile dir is missing! (%s) " % user_profile_dir
                tk.messagebox.showerror("Setup Error", msg)
                raise IOError("Errore Creazione Directory", msg)          
    except:
        raise IOError("Setup dir is missing (2): %s" % msg) 
        
        
    #abs_dir = os.path.join(entry_dir,entry_dir) 
    #print(abs_dir)
    os.chdir(entry_dir)
    
    
    try:
        response_dir = os.path.join(".", data['RESPONSE_DIR'])
        print(response_dir)
        if os.path.isdir(response_dir) == False:
            os.mkdir(response_dir, 755);
        
        save_dir = os.path.join(".", data['SAVE_DIR'])
        print(save_dir)
        if os.path.isdir(save_dir) == False:
            os.mkdir(save_dir, 755);
        
        export_dir = os.path.join(".", data['EXPORT_DIR'])
        print(export_dir)
        if os.path.isdir(export_dir) == False:
            os.mkdir(export_dir, 755);

        img_dir = os.path.join(".", data['IMG_DIR'])
        print(img_dir)
        if os.path.isdir(img_dir) == False:
            os.mkdir(img_dir, 755);
        
        #sh_copy(data['CHECKED'], os.path.join(".","img",data['CHECKED']))
        #sh_copy(data['UNCHECKED'], os.path.join(".","img",data['UNCHECKED']))
        
            
        logs_dir = os.path.join(".", data['LOGS_DIR'])
        print(logs_dir)
        if os.path.isdir(logs_dir) == False:
            os.mkdir(logs_dir, 755);
            
        db_dir = os.path.join(".", data['DB_DIR'])
        print(db_dir)
        if os.path.isdir(db_dir) == False:
            os.mkdir(db_dir, 755);

        funchecked = os.path.join(entry_dir, 'img', data['UNCHECKED'])
        fchecked = os.path.join(entry_dir, 'img', data['CHECKED'])
        im_checked = ImageTk.PhotoImage(Image.open(fchecked))
        im_unchecked = ImageTk.PhotoImage(Image.open(funchecked))
                        
    except IOError as e:
        msg = "Error creating setup dir img (%s), responses (%s), save (%s), export (%s) in %s" % (img_dir, response_dir, save_dir, export_dir, user_profile_dir)
        msg = e #"Error creating setup dir img (%s), responses (%s), save (%s), export (%s) in %s" % (img_dir, response_dir, save_dir, export_dir, user_profile_dir)
        tk.messagebox.showerror("Setup Error", msg)
        raise IOError("Error creating setup dirs", msg)          
            
    logName = os.path.join(logs_dir, "responses.log")
    logging.basicConfig(filename=logName, level=logging.DEBUG, format=FORMAT)




def setup_connection():
    global dbFile, conn, cur, db_dir
    dbFile = os.path.join(db_dir, data['DBNAME'])
    conn = sqlite3.connect(dbFile)
    cur = conn.cursor()
    
    comm = """CREATE TABLE IF NOT EXISTS pdf2ws0f 
                               (id       INTEGER PRIMARY KEY, 
                                filename TEXT, 
                                size     VARCHAR(10),
                                dt_snd   TEXT,
                                dt_rcv   TEXT,
                                status   INTEGER)
           """
    cur.execute(comm)

    comm = """CREATE TABLE IF NOT EXISTS pdf2file0f 
                               (id       INTEGER, 
                                filename VARCHAR(250),
                                size     INTEGER, 
                                data     TEXT,
                                FOREIGN KEY(id) REFERENCES pdf2ws0f(id))
           """
    cur.execute(comm)


class CustomDateEntry(DateEntry):

    def _select(self, event=None):
        date = self._calendar.selection_get()
        if date is not None:
            self._set_text(date.strftime('%m/%d/%Y'))
            self.event_generate('<<DateEntrySelected>>')
        self._top_cal.withdraw()
        if 'readonly' not in self.state():
            self.focus_set()



class MyDateEntry(DateEntry):
    def _validate_date(self):
        if not self.get():
            return True # IMPORTANT!!! Validation must return True/False otherwise it is turned off by tkinter engine
        
        return super()._validate_date()



class zDateEntry(DateEntry):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        
        if self.get_date() == None or self.get_date() == "":
            return None
            
    def get_date(self):
        print("date vale")
        print(self.get())
        if not self.get():
            print("return none")
            return None
        self._validate_date()
        print("prima parse")
        return self.parse_date(self.get())       



"""
DateEntry(scroll_frame, width = 25, background = 'LightCyan3',
                                             foreground ='white',borderwidth=2)
"""



                    
def search_files(qfile, qdatada, qdataa, qstatus):
    qfile = qfile.get().lower()
    qdatada = qdatada.get()
    qdataa = qdataa.get() #.replace("/")
    qstatus = qstatus.get()
    
    print(qdatada)
    
    if qstatus != "":
        try:
            qstatus = int(qstatus)
        except ValueError:
            print("valueerror")
            qstatus=0

        
    query=f""" SELECT id, filename, size, dt_snd, dt_rcv, status FROM pdf2ws0f
              WHERE 1=1 
           """
    if qfile != None:
        query += f" AND lower(filename) LIKE '%{qfile}%'"       

    if qstatus != "":
        query += f" AND status == '{qstatus}'"       

    if qdatada != None:
        query += f" AND substr(dt_snd, 1, 10) >= '{qdatada}'"       
        
    if qdataa != None:
        query += f" AND substr(dt_snd, 1, 10) <= '{qdataa}'"       

    query += """ ORDER BY id DESC"""
    
    print(query)
    cur.execute(query)
    rows=cur.fetchall()
    update(rows)


    
def clear():
    populate()




def setup_frames():    
    global trv, checkall_btn, uncheckall_btn, upload_again_btn, excel_btn

    wrapper1 = LabelFrame(root, text="Last Sent Pdf")
    wrapper2 = LabelFrame(root,  text="Upload Pdf")
    wrapper3 = LabelFrame(root,  text="Search for Pdf")
    wrapper4 = LabelFrame(root,  text="Remove entry")
    
    wrapper1.pack(fill=tk.X, expand="yes", padx=20, pady=10)
    wrapper2.pack(fill=tk.X, expand="yes", padx=20, pady=10)
    wrapper3.pack(fill=tk.X, expand="yes", padx=20, pady=10)
    wrapper4.pack(fill=tk.X, expand="yes", padx=20, pady=10)
    
    wrapper2.pack_propagate(0)
    wrapper3.pack_propagate(0)
    wrapper4.pack_propagate(0)
    
    qfile = StringVar()
    qstatus = StringVar()
    qid = StringVar()
        
    trv = ttk.Treeview(wrapper1, columns=(1,2,3,4,5,6))
    
    vsb = ttk.Scrollbar(wrapper1, orient="vertical", command=trv.yview)
    vsb.pack(side='right', fill='y')
    trv.configure(yscrollcommand=vsb.set)
    
    style=ttk.Style()
    
    style.theme_use("default")

    style.map('Treeview', 
              background=[('selected','brown')]
             )

    style.configure('Treeview', 
                    background='#ffffff',
                    foreground='black',
                    fieldbackground='#ffffff'
                   )
    
    trv.tag_configure('checked', image=im_checked)
    trv.tag_configure('unchecked', image=im_unchecked)
    trv.tag_configure('gray', background='#cccccc')
    

    trv.column("#0", minwidth=40, width=40, stretch=0)
    trv.heading('#0', text="")

    trv.column("#1", minwidth=30, width=50, stretch=0, anchor=CENTER)
    trv.heading('#1', text="Id")

    trv.column("#2", minwidth=200, width=300, stretch=1)
    trv.heading('#2', text="File")

    trv.column("#3", minwidth=50, width=60, stretch=0, anchor=CENTER)
    trv.heading('#3', text="Size")

    trv.column("#4", minwidth=150, width=150, stretch=1)
    trv.heading('#4', text="dt_snd")

    trv.column("#5", minwidth=150, width=150, stretch=1)
    trv.heading('#5', text="dt_rcv")

    trv.column("#6", minwidth=50, width=40, stretch=1)
    trv.heading('#6', text="status")

       
    trv.pack_propagate(0)
    trv.pack()
    
    checkall_btn = Button(wrapper1, text="Check all", command=lambda:toggleCheck2("checked"))
    checkall_btn.pack(side=LEFT, anchor="w", padx=10)

    uncheckall_btn = Button(wrapper1, text="Uncheck all", command=lambda:toggleCheck2("unchecked"))
    uncheckall_btn.pack(side=LEFT, anchor="s", padx=10)

    upload_again_btn = Button(wrapper1, text="Upload again", command=upload_again)
    upload_again_btn.pack(side=LEFT, anchor="s", padx=10)

    excel_btn = Button(wrapper1, text="Export", command=write_workBook)
    excel_btn.pack(side=LEFT, anchor="s", padx=10)




    populate()
    #toggleDisableButton()
    #upload_again_btn['state'] = "disabled"
    #checkall_btn['state'] = "disabled"
    #uncheckall_btn['state'] = "disabled"
    
    #trv.bind('<Double 1>', getrow)
    trv.bind('<Button 1>', toggleCheck)


    
    lblf = Label(wrapper2, text="Upload file...")
    lblf.grid(column=0, row=0, padx=5, pady=5)
    
    upload_btn = Button(wrapper2, text="Upload", command=upload_files)
    upload_btn.grid(column=1, row=0, padx=5, pady=5)
    
    
    lblf = Label(wrapper3, text="FileName...")
    lblf.grid(column=0, row=0, padx=5, pady=5)
    entf = Entry(wrapper3, textvariable=qfile)
    entf.grid(column=1, row=0, padx=5, pady=5)
    
    lbls = Label(wrapper3, text="Status.....")
    lbls.grid(column=2, row=0, padx=5, pady=5)
    ents = Entry(wrapper3, textvariable=qstatus)
    ents.grid(column=3, row=0, padx=5, pady=5)
    
    lbld = Label(wrapper3, text="Date from...")
    lbld.grid(column=0, row=1, padx=5, pady=5)
    entda = MyDateEntry(wrapper3,selectmode='day', date_pattern='Y-mm-d')
    entda.grid(column=1, row=1, padx=5, pady=5)
    
    #entda._set_text(entda._date.strftime('%m/%d/%Y'))    
    #entda.configure(validate='none')

    lbld = Label(wrapper3, text="Date to.....")
    lbld.grid(column=2, row=1, padx=5, pady=5)
    enta = MyDateEntry(wrapper3,selectmode='day', date_pattern='Y-mm-d')
    enta.grid(column=3, row=1, padx=5, pady=5)
    
    #enta._set_text(enta._date.strftime('%m/%d/%Y'))    
    
    #enta.configure(validate='none')
    
    #done with lambda in order not to make qfile, entda, enta, qstatus globals
    search_btn = Button(wrapper3, text="Search Files", command=lambda:search_files(qfile, entda, enta, qstatus))
    search_btn.grid(column=0, row=2, padx=5, pady=5)
    clear_btn = Button(wrapper3, text="Clear", command=clear)
    clear_btn.grid(column=1, row=2, padx=5, pady=5)
    
    
    lbldel = Label(wrapper4, text="Id num.....")
    lbldel.grid(column=0, row=0, padx=5, pady=5)
    entdel = Entry(wrapper4, textvariable=qid)
    entdel.grid(column=1, row=0, padx=5, pady=5)
    
    #done with lambda in order not to make qid global
    del_btn = Button(wrapper4, text="Delete id", command=lambda:delete_id(qid))
    del_btn.grid(column=2, row=0, padx=5, pady=5)
    
    exit_btn = Button(wrapper4, text="Exit App", command=root.destroy)
    #exit_btn.grid(padx=5, pady=5)
    exit_btn.pack(side=tk.RIGHT, padx=10)
    

                                    

def upload_files(here_file=None):
    error = False
    save = True
    #os.chdir(entry_dir)
    
    def get_initial_dir():
        try:
            if os.path.isdir(os.path.join(os.environ.get("USERPROFILE"),"download")):
                initial_dir=os.path.join(os.environ.get("USERPROFILE"),"download")
            elif os.path.isdir(os.path.join(os.environ.get("USERPROFILE"),"downloads")):
                initial_dir=os.path.join(os.environ.get("USERPROFILE"),"downloads")
            elif os.path.isdir(os.environ.get("USERPROFILE")):
                initial_dir=os.environ.get("USERPROFILE")
            else:
                initial_dir=os.getcwd()
        except:
            initial_dir=os.getcwd()
            
        return initial_dir
        
            
    if here_file != None:
        print("herefile ", here_file)
        files = (os.path.join(save_dir, here_file),)
        save=False
    else:
        tk_chooseFile = tk.filedialog.askopenfilenames if data['MULTI_FILES'] else filedialog.askopenfilename
        files=tk_chooseFile(initialdir=get_initial_dir(),
                      title="Please select a file to upload:",
                      filetypes=allowedFileTypes)
        
        
        
    print("here file ")
    print(type(files))    

    for filename in files:
        abs_file = os.path.abspath(filename)
        size = os.stat(abs_file).st_size
        file = os.path.basename(filename).lower()
        ext = ''.join(file.split(".")[-1])
        fbase = ''.join(file.split(".")[:-1])
        print(fbase, ext)
        if ext not in data['EXTENSIONS']:
            msg = "File %s cannot be sent due to wrong extension: %s!" % (file, ext)
            tk.messagebox.showerror("File Input Error", msg)
            #win32api.MessageBox(0, msg, "Critical Error", 0x00001000)         
            logging.error(msg) 
            break
            
        
        files_ = {'file': (file, open(filename, 'rb'), 'application/pdf', {'Expires': '0'})}
        dt_snd = get_timestamp()
        r = requests.post(data['URL'], files=files_)
        dt_rcv = get_timestamp()
        
        file_response = os.path.join(response_dir, "%s%s" % (fbase, ".response"))

        file_save = os.path.join(save_dir, file)
        print("---")
        print(file_save)
        
        
        fh_resp = open(file_response,"w")
        
        fh_resp.write("status code: %s" % str(r.status_code))
        fh_resp.write(str(r.headers))
        fh_resp.write(r.text)
        fh_resp.close()
    
        id = record_upload(os.path.basename(filename.lower()), size, dt_snd, dt_rcv, r.status_code)
            
        if r.status_code not in CODES_OK:
            print("ecco status code %s" %  r.status_code )
            msg = "Upload failed for File %s. status_code: %s!" % (file, r.status_code)
            tk.messagebox.showerror("Failed file upload", msg)
            #win32api.MessageBox(0, msg, "Critical Error", 0x00001000) 
            logging.error(msg) 
            r.raise_for_status()
        else:
            msg = "Upload Successfull for File %s. status_code: %s!" % (file, r.status_code)
            tk.messagebox.showinfo("File upload successfull", msg)
            #win32api.MessageBox(0, msg, "Critical Error", 0x00001000) 
            logging.error(msg) 
            
            if SAVE_FILE_DIR and save:
                try:
                    sh_copy(abs_file, file_save)
                except:
                    msg = "Error saving file %s!" % os.path.basename(file)
                    logging.error(msg) 
                    tk.messagebox.showerror("Saving Error", msg)
            
            if SAVE_FILE_DB:
                try:
                    file_load(id, abs_file)
                except (e):
                    msg = "Error saving file %s into db. Error %s!" % (os.path.basename(file), str(e))
                    logging.error(msg) 
                    tk.messagebox.showerror("Saving Error", msg)
            
            
    

def get_timestamp():
    return dt.now().strftime(format='%Y-%m-%d %H:%M:%S.%f')[:-3]
    

    
def update(rows):
    trv.delete(*trv.get_children())
    for i, row in enumerate(rows):
        tags=['unchecked'] if i%2 == 1 else ['unchecked','gray']
        print("inserted %d %s" % (i, str(tags)))
        trv.insert('', 'end', values=row, tags=tags)
    
    toggleDisableButton()
   

def toggleCheck2(opt):
    for rowid in trv.get_children():
        trv.item(rowid, tags=opt)
    toggleDisableButton()
    refresh_stripes()


def countChecked(opt="checked"):
    cnt=0
    for rowid in trv.get_children():
        #iid = trv.index(item)
        if trv.item(rowid, "tags")[0] == opt:
            cnt +=1
            
    print("checked: ",cnt)
    return cnt


def getCheckedIds(opt="checked"):
    cnt=0
    for rowid in list(reversed(trv.get_children())):
        print("getcheckedids")
        print(type(trv.item(rowid, "values")))
        if trv.item(rowid, "tags")[0] == opt:
            yield trv.item(rowid)
    
    
def toggleDisableButton():
    print(trv.get_children())
    tot = len(list(trv.get_children()))
    print("tot %s" % tot)
    if tot:
        excel_btn['state'] = "active"
        if countChecked() == tot:
            checkall_btn['state'] = "disabled"
            uncheckall_btn['state'] = "active"
            upload_again_btn['state'] = "active"
        elif countChecked() == 0:    
            uncheckall_btn['state'] = "disabled"
            checkall_btn['state'] = "active"
            upload_again_btn['state'] = "disabled"
        elif countChecked():     
            uncheckall_btn['state'] = "active"
            checkall_btn['state'] = "active"
            upload_again_btn['state'] = "active"
    else:
        checkall_btn['state'] = "disabled"
        uncheckall_btn['state'] = "disabled"
        upload_again_btn['state'] = "disabled"
        excel_btn['state'] = "disabled"
            
            
    
        
def toggleCheck(event):
    rowid = trv.identify_row(event.y)
    if rowid != None and rowid != "":
        tag = trv.item(rowid, "tags")[0]
        #print("--- ", rowid)
        #print(tag)
        #tags = list(trv.item(rowid, "tags"))
        #tags.remove(tag)
        #trv.item(rowid, tags=tags)
        if tag == "checked":
            trv.item(rowid, tags="unchecked")
        else:
            trv.item(rowid, tags="checked")
        toggleDisableButton()
        refresh_stripes(rowid)


def refresh_stripes(rowid=-1):
    for i, rid in enumerate(trv.get_children()):
        if rid != rowid:
            if i%2 == 0:
                tags = (trv.item(rid, "tags")[0],"gray")
            else:
                tags = (trv.item(rid, "tags")[0],)
            
            trv.item(rid, tags=tags)



def upload_again():
    """
    #for row in trv.get_children():
    #for row in trv_list:
        item = trv.item(row)
        print("qwe---")
        print(item["tags"])
        print(item)
        print(item['values'][0])
        print(item['values'][1])    
        if item["tags"][0] == "checked":
            upload_files(item['values'][1])
    """
    trv_list = [ trv.item(row) for row in trv.get_children() ]
    for item in trv_list:
        #item = trv_list(row)
        print("qwe---")
        print(item)
        print(item["tags"])
        print(item)
        print(item['values'][0])
        print(item['values'][1])    
        if item["tags"][0] == "checked":
            upload_files(item['values'][1])
        

def populate():
    query="SELECT id, filename, size, dt_snd, dt_rcv, status FROM pdf2ws0f ORDER BY id DESC"
    cur.execute(query)
    rows=cur.fetchall()
    update(rows)


def record_upload(filename, size, dt_snd, dt_rcv, status):
    if filename == "" or \
       size == "" or \
       dt_snd == "" or \
       dt_rcv == "" or \
       status == "":
       return 
    
    #print(filename, dt_snd, dt_rcv, status)
    try:       
        cur.execute('INSERT INTO pdf2ws0f (filename, size, dt_snd, dt_rcv, status) VALUES (?,?,?,?,?)', (filename, rs(size), dt_snd, dt_rcv, status))
        rows=get_affected_rows()
        rowid = cur.lastrowid
        print(type(rows))
        print(rows)
        if rows != 1:   
            msg="Inserted rows: %d" % rows
            logging.error(msg) 
            messagebox.showerror("Insert Error", msg) 
            conn.rollback()
            logging.error("rollback occurred") 
            return True
    except sqlite3.Error as e:
        conn.rollback()
        messagebox.showerror("Insert Error", str(e))
    else:
        conn.commit()
    finally:
        populate()
        return rowid
        
    
def file_load(id, file):
    
    error = False
    print("foreing")
    print(id, file)
    basename = os.path.basename(file)
    text_data = base64.b64encode(open(file,"rb").read())
    size = len(text_data)
    
    try:       
        #cur.execute("INSERT INTO pdf2file0f (id, filename, data) VALUES ({0},{1},{2})".format(id, basename, text_data))
        sql="INSERT INTO pdf2file0f (id, filename, size, data) VALUES (?,?,?,?) "
        cur.execute(sql, (id, basename, size, text_data))
    except sqlite3.Error as e:
        print("error sql")
        error=True
        msg="Error inserting pdf2file0f filename %s, error: %s" % (basefile, str(e))
        messagebox.showerror(msg)
        print(msg)
        logging.error(msg) 
        conn.rollback()

    except e:
        conn.rollback()
        print("error sql2")
        error=True
        msg="Error inserting pdf2file0f filename %s, error: %s" % (basefile, str(e))
        logging.error(msg) 
        messagebox.showerror(msg)
    else:
        print("ok sql")
        error=False
        conn.commit()
    finally:
        print("return sql2")
        return error
        
    
    
def getrow():
    #rowid = trv.identify_row(event.y)
    item = trv.item(trv.focus())
    logging.info("%s %s %s %s %s" % (item['values'][0], item['values'][1], item['values'][2], item['values'][3], item['values'][4]))
     
    #t1.set(item['values'][0]) t2.set(item['values'][1]) t3.set(item['values'][2]) t4.set(item['values'][3]) t5.set(item['values'][4])
    
    
def get_affected_rows():
    query = "SELECT changes()"
    cur.execute(query)
    return cur.fetchone()[0]
       

def delete_id(qid):

    tot = len(list(trv.get_children()))
    if tot == 0:
        return False
    
    try:
        id=int(qid.get())
    except ValueError:
        id=0    
    
    cntCkd = countChecked()
    
    if cntCkd == 0 and id == 0:   
        return False
        

    if cntCkd == tot and tot > 1: 
        msg="Are you sure you want to delete all %d rows?" % tot
        ids = (row['values'][0] for row in getCheckedIds())
    elif cntCkd > 0:    
        ids = [row['values'][0] for row in getCheckedIds()]
        msg="Are you sure you want to delete these ids? (%s) " % ', '.join(str(id) for id in ids)
    else:
        msg="Are you sure you want to delete this id %d?" % id
        ids = [id]
    if messagebox.askyesno("Confirm please", msg) == False:
        return False


    try:
        for id in ids:
            print("to delete ", id)
            
            query=""" SELECT id FROM pdf2ws0f WHERE id = ? """ 
            cur.execute(query, (id,))
            
            if cur.fetchone() != None:
                query=""" SELECT id FROM pdf2file0f WHERE id = ? """ 
                cur.execute(query, (id,))
                if cur.fetchone() != None:
                    query=""" DELETE FROM pdf2file0f WHERE id = ? """ 
                    deleted = cur.execute(query, (id,)).rowcount
                    if deleted != 1:     
                    #if get_affected_rows() != 1:   
                        msg="Error deleting rows from pdf2file0f. Deleted rows: %d" % deleted
                        logging.error(msg)
                        messagebox.showerror("Delete Error", msg) 
                        conn.rollback()
                        logging.error("rollback occurred") 
                        return True
                    logging.info("deleted id %d from pdf2file0f" % id) 
                
                query=""" DELETE FROM pdf2ws0f WHERE id = ? """ 
                deleted = cur.execute(query, (id,)).rowcount
                if deleted != 1:     
                #if get_affected_rows() != 1:   
                    msg="Error deleting rows from pdf2ws0f. Deleted rows: %d" % deleted
                    logging.error(msg)
                    messagebox.showerror("Delete Error", msg) 
                    conn.rollback()
                    logging.error("rollback occurred") 
                    return True
                logging.info("deleted id %d from pdf2ws0f" % id) 


    except sqlite3.Error as e:
        conn.rollback()
        logging.error("rollback occurred") 
        messagebox.showerror("Insert Error", str(e))
    else:
        conn.commit()
        populate()
        
    return True            
 

def write_workBook():
    bstart=True
    
    column_width=(None, 5, 45, 8, 25, 25, 8)
    
    font_header = Font(name='Calibri',
                 size=13,
                 bold=True,
                 color='00000000')
    
    fillColorHeader = PatternFill(start_color="FFFF00", 
                                  end_color="FFFF00", 
                                  fill_type = "solid")
    
    fillColorStriped = PatternFill(start_color="CCCCCC", 
                                  end_color="CCCCCC", 
                                  fill_type = "solid")    
    #print("before wotkbook")
    #print(trv.heading('#1'))
    for i, rowid in enumerate(trv.get_children()):
        if bstart:
            wb = Workbook()
            ws1 = wb.active
            ws1.title = "uploaded files"

        for j, column in enumerate(trv.item(rowid, "values")):
            if bstart:
                ws1.column_dimensions[get_column_letter(j+1)].width = column_width[j+1]
                wcell = ws1.cell(column=j+1, row=i+1, value="{0}".format(trv.heading('#%s' % str(j+1))['text'] ))
                wcell.font = font_header
                wcell.fill = fillColorHeader
            
            wcell_body = ws1.cell(column=j+1, row=i+2, value="{0}".format(column))
            if i>0 and i%2 == 1:
                wcell_body.fill = fillColorStriped
                 
        bstart=False
    


    if not bstart: 
        try:
            exp_file = os.path.join(export_dir, 'file_sent_%s.xlsx' % dt.strftime(dt.now(),'%Y%m%d'))
            wb.save(filename=exp_file)
        except IOError as e:
            messagebox.showerror("Write File Error", "Unable to write file %s! Check permission or if it is already open! %s" % (os.path.basename(exp_file), str(e)))     
        else:
            messagebox.showinfo("Write File Info", "Created file %s!" % os.path.basename(exp_file))     



if __name__ == "__main__":

    try:
        setup_profile()  
    except IOError as e:
        msg = "Aborted! creating user profile! Error: %s" % str(e)  
        logging.critical(msg)  
        tk.messagebox.showerror("Setup Error", msg) 
        exit()
    
    try:
        setup_connection()  
    except sqlite3.Error as e:
        msg = "Aborted! Error connecting to db! Error: %s" % str(e)  
        logging.critical(msg)  
        tk.messagebox.showerror("Setup Error", msg) 
        exit()

 
    setup_frames()
    root.mainloop()






